from __future__ import annotations

import io

from .optimizer import _load_excel_from_s3


def export_trucks_workbook(s3_key: str, sheet_name: str | None):
    # TODO: Build real workbook (Truck Summary + Order Details)
    # For now, return minimal empty workbook to validate plumbing
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Truck Summary")
    else:
        ws.title = "Truck Summary"
    ws.append(["Truck Number", "Customer", "City",
              "State", "Total Weight"])  # headers

    # Return as stream
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


essential_dh_sheets = ["Late+NearDue", "WithinWindow"]


def export_dh_load_list_workbook(s3_key: str, sheet_name: str | None):
    # TODO: Implement exact-format DH Load List per sample
    import openpyxl

    wb = openpyxl.Workbook()
    # Ensure first default sheet is correct name
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(essential_dh_sheets[0])
    else:
        ws.title = essential_dh_sheets[0]
    # Add second sheet
    wb.create_sheet(essential_dh_sheets[1])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
